# api/views/cn/export_mentors.py
"""
GET /api/cn/mentors/export-csv/
  ?file_format=csv|xlsx   (défaut : xlsx)

Export national (CN) : tous les mentors de tous les pôles, en Excel ou CSV.
"""
import csv
import io

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from core.models import Mentor
from api.permissions import IsCN

HEADERS = [
    'Code pôle', 'Pôle', 'ID',
    'Prénom', 'Nom', 'Email', 'Téléphone',
    'Ville', 'Code postal', 'Département',
    'Association',
    'Formé', 'Date formation',
    'Statut', 'Disponibilité', 'Capacité max',
    'Mentorats actifs', 'Mentorats terminés',
    'Particularité pour l\'affectation',
]


def _build_rows(mentors):
    rows = []
    for m in mentors:
        rows.append([
            m.pole.code if m.pole_id else '',
            m.pole.name if m.pole_id else '',
            m.id,
            m.first_name,
            m.last_name,
            m.email,
            m.phone or '',
            m.city or '',
            m.code_postal or '',
            m.department.name if m.department_id else '',
            m.association.name if m.association_id else '',
            'Oui' if m.is_trained else 'Non',
            m.training_date.strftime('%d/%m/%Y') if m.training_date else '',
            'Actif' if m.is_active else 'Inactif',
            m.disponibilite_reelle,
            m.max_capacity,
            getattr(m, 'nb_actifs', 0),
            getattr(m, 'nb_termines', 0),
            m.observations or '',
        ])
    return rows


class ExportMentorsNationalCsvView(APIView):
    permission_classes = [IsAuthenticated, IsCN]

    def get(self, request):
        from django.db.models import Count, Q

        mentors = (
            Mentor.objects
            .select_related('pole', 'association', 'department')
            .annotate(
                nb_actifs=Count('mentorats', filter=Q(mentorats__status='ACTIVE')),
                nb_termines=Count('mentorats', filter=Q(mentorats__status__in=['CLOSED', 'ABORTED'])),
            )
            .order_by('pole__code', 'last_name', 'first_name')
        )

        rows = _build_rows(mentors)
        base_name = "mentors_national"

        fmt = request.query_params.get('file_format', 'xlsx').lower()
        if fmt == 'csv':
            return self._csv_response(rows, base_name)
        return self._xlsx_response(rows, base_name)

    def _csv_response(self, rows, base_name):
        filename = base_name + '.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('﻿')
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow(row)
        return response

    def _xlsx_response(self, rows, base_name):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mentors"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1E4A8A")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 32

        for row in rows:
            ws.append(row)

        col_widths = {
            1: 12, 2: 20, 3: 8,
            4: 14, 5: 16, 6: 28, 7: 14,
            8: 18, 9: 10, 10: 16,
            11: 14,
            12: 8, 13: 14,
            14: 10, 15: 14, 16: 14,
            17: 14, 18: 16,
            19: 36,
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = base_name + '.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
