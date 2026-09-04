# api/views/cn/export_animateurs.py
"""
GET /api/cn/animateurs/export-csv/
  ?file_format=csv|xlsx   (défaut : xlsx)

Export national (CN) : tous les animateurs de tous les pôles, en Excel ou CSV.
"""
import csv
import io

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from core.models import Animateur
from api.permissions import IsCN

HEADERS = [
    'ID', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Ville',
    'Code pôle', 'Pôle', 'Association',
    'Rôle', 'Statut', 'Compte créé',
]


def _role_label(a):
    if a.is_acp:
        return 'APC'
    if a.is_ap:
        return 'AP'
    return 'Sans rôle'


def _build_rows(animateurs):
    rows = []
    for a in animateurs:
        rows.append([
            a.id,
            a.first_name,
            a.last_name,
            a.email,
            a.phone or '',
            a.city or '',
            a.pole.code if a.pole_id else '',
            a.pole.name if a.pole_id else '',
            a.association.name if a.association_id else '',
            _role_label(a),
            'Actif' if a.is_active else 'Inactif',
            'Oui' if a.user_id else 'Non',
        ])
    return rows


class ExportAnimateursNationalCsvView(APIView):
    permission_classes = [IsAuthenticated, IsCN]

    def get(self, request):
        animateurs = (
            Animateur.objects
            .select_related('pole', 'association')
            .order_by('pole__code', 'last_name', 'first_name')
        )

        rows = _build_rows(animateurs)
        base_name = "animateurs_national"

        fmt = request.query_params.get('file_format', 'xlsx').lower()
        if fmt == 'csv':
            return self._csv_response(rows, base_name)
        return self._xlsx_response(rows, base_name)

    def _csv_response(self, rows, base_name):
        filename = base_name + '.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('﻿')
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow(row)
        return response

    def _xlsx_response(self, rows, base_name):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Animateurs"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1E4A8A")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 32

        for row in rows:
            ws.append(row)

        col_widths = {
            1: 8, 2: 14, 3: 16, 4: 28, 5: 14, 6: 18,
            7: 12, 8: 20, 9: 14,
            10: 12, 11: 10, 12: 12,
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = base_name + '.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
