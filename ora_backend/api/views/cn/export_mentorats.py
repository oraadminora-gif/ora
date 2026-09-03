# api/views/cn/export_mentorats.py
"""
GET /api/cn/mentorats/export-csv/
  ?date_debut=YYYY-MM-DD (optionnel — filtre sur assigned_at)
  ?date_fin=YYYY-MM-DD    (optionnel)
  ?file_format=csv|xlsx   (défaut : xlsx)

Export national (CN) : mêmes colonnes que l'export pôle
(api/views/pole/export_csv.py), avec une colonne « Pôle » en plus puisqu'il
regroupe les mentorats de tous les pôles du programme.
"""
import csv
import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from core.models import Mentorat
from api.permissions import IsCN
from api.views.pole.export_csv import HEADERS, _build_rows

_ID_MENTORAT_IDX = HEADERS.index('ID Mentorat')

# « Pôle » puis « ID Mentorat » en tête, pour identifier chaque ligne au
# premier coup d'œil dans cet export qui regroupe tous les pôles.
HEADERS_NATIONAL = ['Pôle', 'ID Mentorat'] + [
    h for i, h in enumerate(HEADERS) if i != _ID_MENTORAT_IDX
]


class ExportMentoratsNationalCsvView(APIView):
    permission_classes = [IsAuthenticated, IsCN]

    def get(self, request):
        qs = (
            Mentorat.objects
            .select_related(
                'pole',
                'mentor', 'mentor__association',
                'young_request', 'young_request__etablissement',
                'ap_responsable', 'ap_responsable__association',
            )
            .prefetch_related('financements__financement', 'suivis')
            .order_by('pole__code', 'assigned_at')
        )

        # ── Filtres plage de dates ────────────────────────────────
        date_debut = request.query_params.get('date_debut')
        date_fin   = request.query_params.get('date_fin')
        if date_debut:
            try:
                d = datetime.strptime(date_debut, '%Y-%m-%d').date()
                qs = qs.filter(assigned_at__gte=d)
            except ValueError:
                pass
        if date_fin:
            try:
                d = datetime.strptime(date_fin, '%Y-%m-%d').date()
                qs = qs.filter(assigned_at__lte=d)
            except ValueError:
                pass

        base_name = "mentorats_national"
        if date_debut:
            base_name += f"_{date_debut}"
        if date_fin:
            base_name += f"_au_{date_fin}"

        mentorats = list(qs)
        rows = []
        for m, row in zip(mentorats, _build_rows(mentorats)):
            pole_label = f"{m.pole.code} — {m.pole.name}" if m.pole_id else ''
            id_mentorat = row[_ID_MENTORAT_IDX]
            reste = [v for i, v in enumerate(row) if i != _ID_MENTORAT_IDX]
            rows.append([pole_label, id_mentorat] + reste)

        fmt = request.query_params.get('file_format', 'xlsx').lower()
        if fmt == 'csv':
            return self._csv_response(rows, base_name)
        return self._xlsx_response(rows, base_name)

    # ── CSV ───────────────────────────────────────────────────────
    def _csv_response(self, rows, base_name):
        filename = base_name + '.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('﻿')  # UTF-8 BOM pour Excel

        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writerow(HEADERS_NATIONAL)
        for row in rows:
            writer.writerow(row)
        return response

    # ── Excel (xlsx) ──────────────────────────────────────────────
    def _xlsx_response(self, rows, base_name):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mentorats"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1E4A8A")  # bleu ORA
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(HEADERS_NATIONAL)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 36

        for row in rows:
            ws.append(row)

        ws.column_dimensions['A'].width = 18  # Pôle
        for col_idx in range(2, len(HEADERS_NATIONAL) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = base_name + '.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
