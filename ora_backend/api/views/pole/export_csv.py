# api/views/pole/export_csv.py
"""
GET /api/pole/mentorats/export-csv/
  ?date_debut=YYYY-MM-DD  (optionnel — filtre sur assigned_at)
  ?date_fin=YYYY-MM-DD    (optionnel)
  ?format=csv|xlsx        (défaut : csv)

Retourne un fichier CSV (UTF-8 BOM) ou Excel avec une ligne par mentorat.
Ordre des colonnes : Jeune → Mentor → AP → Mentorat → Financements → Suivis
"""
import csv
import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from core.models import Mentorat
from api.permissions import IsACP

PROBLEMATIQUES_LABELS = {
    'aide_informatique':  'Aide informatique',
    'fle':                'Apprentissage du français (FLE)',
    'changer_employeur':  "Changer d'employeur",
    'handicap':           'Handicap',
    'logement':           'Logement',
    'orientation':        'Orientation',
    'prob_administratif': 'Problème administratif',
    'prob_financier':     'Problème financier — Gérer Budget',
    'fragilite_mentale':  'Fragilité mentale',
    'prep_dossier':       'Prép dossier professionnel',
    'relation_employeur': "Relation avec l'employeur",
    'recherche_contrat':  'Recherche contrat apprentissage',
    'salaire':            'Salaire / Respect de convention',
    'soutien_moral':      'Soutien moral',
    'soutien_scolaire':   'Soutien scolaire',
    'autre':              'Autre',
}

# ── Ordre : Jeune → Mentor → AP → Mentorat → Financements → Suivis ──────────
HEADERS = [
    # ── Jeune ─────────────────────────────────────────────────────
    'Prénom jeune', 'Nom jeune', 'Email jeune', 'Tél jeune',
    'Date naissance', 'Genre', 'Commune', 'Code postal jeune',
    'Diplôme préparé', 'Situation', 'Urgence (1-5)',
    'Nom établissement',
    # ── Mentor ────────────────────────────────────────────────────
    'Prénom mentor', 'Nom mentor', 'Email mentor', 'Tél mentor',
    'Ville mentor', 'CP mentor', 'Association mentor',
    'Formé', 'Date formation',
    # ── AP ────────────────────────────────────────────────────────
    'AP responsable', 'Association AP',
    # ── Mentorat ──────────────────────────────────────────────────
    'ID Mentorat', 'Statut', 'Date début', 'Date fin prévue', 'Date clôture',
    'Alerte rouge', 'Dernier contact', 'Raison clôture', 'Notes de suivi',
    'Problématiques',
    # ── Financements ──────────────────────────────────────────────
    'Financeurs (nom | code dossier)',
    # ── Suivis ────────────────────────────────────────────────────
    'Nb rencontres', 'Durée totale (min)',
    'Types de rencontres', 'Dernier suivi',
]

STATUS_LABELS = {
    'PENDING': 'En attente',
    'ACTIVE':  'Actif',
    'CLOSED':  'Clôturé',
    'ABORTED': 'Abandonné',
}


def _fmt(val):
    """Convertit None/bool en chaîne propre."""
    if val is None:
        return ''
    if isinstance(val, bool):
        return 'Oui' if val else 'Non'
    return str(val)


def _build_rows(qs):
    """Génère les lignes de données (liste de listes) pour un queryset de mentorats."""
    rows = []
    for m in qs:
        req    = m.young_request
        mentor = m.mentor

        # Problématiques
        prob_labels = ' | '.join(
            PROBLEMATIQUES_LABELS.get(c, c)
            for c in (m.problematiques or [])
        )

        # Etablissement
        nom_etab = (
            req.etablissement.nom if req.etablissement_id else req.nom_etablissement
        ) or ''

        # Financements
        fins = m.financements.all()
        financeurs_str = ' | '.join(
            f"{mf.financement.nom}"
            + (f" [{mf.code_specifique}]" if mf.code_specifique else "")
            for mf in fins
        )

        # Suivis
        suivis = list(m.suivis.all())
        nb_suivis   = len(suivis)
        duree_total = sum(s.duree_minutes for s in suivis)
        types_set   = sorted({s.get_type_rencontre_display() for s in suivis})
        types_str   = ' | '.join(types_set)
        dernier     = suivis[0].date_rencontre.isoformat() if suivis else ''

        rows.append([
            # Jeune
            req.first_name,
            req.last_name,
            req.email or '',
            req.phone or '',
            _fmt(req.birth_date),
            req.get_gender_display() if req.gender else '',
            getattr(req, 'commune', '') or '',
            getattr(req, 'code_postal', '') or '',
            req.get_diplome_prepare_display() if req.diplome_prepare else '',
            req.get_situation_display() if req.situation else '',
            nom_etab,
            # Mentor
            mentor.first_name,
            mentor.last_name,
            mentor.email,
            mentor.phone or '',
            mentor.city or '',
            mentor.code_postal or '',
            mentor.association.name,
            'Oui' if mentor.is_trained else 'Non',
            _fmt(mentor.training_date),
            # AP
            (
                f"{m.ap_responsable.first_name} {m.ap_responsable.last_name}".strip()
                if m.ap_responsable else ''
            ),
            m.ap_responsable.association.name if m.ap_responsable else '',
            # Mentorat
            m.id,
            STATUS_LABELS.get(m.status, m.status),
            _fmt(m.assigned_at),
            _fmt(m.expected_end_date),
            _fmt(m.closed_at),
            _fmt(m.alerte_rouge),
            _fmt(m.dernier_contact),
            m.closure_reason or '',
            (m.notes_suivi or '').replace('\n', ' '),
            prob_labels,
            # Financements
            financeurs_str,
            # Suivis
            nb_suivis,
            duree_total,
            types_str,
            dernier,
        ])
    return rows


class ExportMentoratsCsvView(APIView):
    permission_classes = [IsAuthenticated, IsACP]

    def get(self, request):
        if not hasattr(request.user, 'animateur'):
            return HttpResponse('Pas de pôle', status=400)
        pole_id = request.user.animateur.pole_id

        qs = (
            Mentorat.objects
            .filter(pole_id=pole_id)
            .select_related(
                'mentor', 'mentor__association',
                'young_request', 'young_request__etablissement',
                'ap_responsable', 'ap_responsable__association',
            )
            .prefetch_related(
                'financements__financement',
                'suivis',
            )
            .order_by('assigned_at')
        )

        # ── Filtres plage de dates ────────────────────────────────
        date_debut = request.query_params.get('date_debut')
        date_fin   = request.query_params.get('date_fin')
        if date_debut:
            try:
                d = datetime.strptime(date_debut, '%Y-%m-%d').date()
                qs = qs.filter(assigned_at__gte=d)
            except ValueError:
                pass
        if date_fin:
            try:
                d = datetime.strptime(date_fin, '%Y-%m-%d').date()
                qs = qs.filter(assigned_at__lte=d)
            except ValueError:
                pass

        # ── Nom de fichier de base ────────────────────────────────
        base_name = f"mentorats_{pole_id}"
        if date_debut:
            base_name += f"_{date_debut}"
        if date_fin:
            base_name += f"_au_{date_fin}"

        fmt = request.query_params.get('file_format', 'csv').lower()

        if fmt == 'xlsx':
            return self._xlsx_response(qs, base_name)
        return self._csv_response(qs, base_name)

    # ── CSV ───────────────────────────────────────────────────────
    def _csv_response(self, qs, base_name):
        filename = base_name + '.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff')  # UTF-8 BOM pour Excel

        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
        writer.writerow(HEADERS)
        for row in _build_rows(qs):
            writer.writerow(row)
        return response

    # ── Excel (xlsx) ──────────────────────────────────────────────
    def _xlsx_response(self, qs, base_name):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mentorats"

        # En-têtes stylisés
        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="1E4A8A")  # bleu ORA
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 36

        # Données
        for row in _build_rows(qs):
            ws.append(row)

        # Largeurs automatiques (approximation)
        col_widths = {
            1: 14, 2: 16, 3: 28, 4: 14,   # Jeune
            5: 14, 6: 10, 7: 18, 8: 10,
            9: 22, 10: 22, 11: 10, 12: 24,
            13: 14, 14: 16, 15: 28, 16: 14,  # Mentor
            17: 16, 18: 10, 19: 14, 20: 6, 21: 14,
            22: 22, 23: 14,                  # AP
            24: 10, 25: 12, 26: 14, 27: 14, 28: 14,  # Mentorat
            29: 10, 30: 14, 31: 20, 32: 40, 33: 30,
            34: 30,                          # Financements
            35: 12, 36: 14, 37: 24, 38: 14, # Suivis
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Figer la première ligne
        ws.freeze_panes = "A2"

        # Écriture en mémoire
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = base_name + '.xlsx'
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
